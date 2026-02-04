import logging
import time
import openpyxl
import csv

from alive_progress import alive_bar
from pathlib import Path
from tqdm import tqdm


class ExcelToCsv:
    """该类将excel文件转换为csv文件"""

    def __init__(self, project_name: str = "ExcelToCsv", method: str = "dir"):
        """初始化 ExcelToCsv 类实例

        Args:
            project_name (str, optional): 项目名称. Defaults to "ExcelToCsv".
            method (str, optional): 该类的转换模式, 可选的值有: "dir", "file". Defaults to "dir".
        """
        self.logger: logging.Logger = logging.getLogger(f"{project_name}.{__name__}")
        self.method: str | None = self.__verify_params(method)

    def __verify_params(self, method: str) -> str | None:
        """检验类的初始化参数是否正确"""

        method_list: list[str] = ["dir", "file"]
        if method not in method_list:
            self.logger.error(
                f"ExcelToCsv类的method参数没有{method}值, method参数值有: 'dir', 'file'."
            )
            raise ValueError()
        else:
            return method

    def path_exists(self, file_path: str) -> Path:
        """判断输入的文件路径是否符合规范或存在

        Args:
            path (str): 文件/文件夹路径

        Returns:
            Path: 路径的Path对象
        """

        try:
            path: Path = Path(file_path)
        except (TypeError, ValueError) as e:
            self.logger.error(e, f"输入的路径: {file_path} 不符合规范, 请重新输入")
            raise e

        if self.method == "dir":
            if path.is_file():
                self.logger.error(f"输入的路径: {path} 不是文件夹路径, 请重新输入")
                raise ValueError()

        elif self.method == "file":
            if path.is_dir():
                self.logger.error(f"输入的路径: {path} 不是文件夹路径, 请重新输入")
                raise ValueError()

        return path

    def to_csv(
        self,
        path: Path,
        output_dir: Path | None = None,
        encoding: str = "utf-8",
    ) -> Path | None:
        """读取Excel文件, 将文件转化为csv文件

        Args:
            path (Path): 带转换的文件路径.
            encoding (str, optional): 文件的编码格式. Defaults to "utf-8".
            index (bool, optional): 是否保留索引. Defaults to False.

        Returns:
            Path | None: csv文件lujing或者空值.
        """

        if path.suffix == ".csv":
            self.logger.info(f"\n{path.name}已经是csv文件, 无需转换!")
            return

        if self.method == "file":
            csv_path = path.with_suffix(".csv")
        elif self.method == "dir":
            if output_dir == None:
                self.logger.info(
                    f"{self.method}模式下, to_csv方法output_dir参数值不能为None."
                )
                raise ValueError()
            csv_path = output_dir / f"{path.stem}.csv"

        self.logger.info(f"\n--读取 '{path.name}' --")
        with alive_bar(title=f"读取 '{path.name}' ", spinner="waves") as bar:
            start_time: float = time.time()

            wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
            ws = wb.active

            elapsed = time.time() - start_time

        self.logger.info(
            f"读取完成! 耗时: {elapsed:.2f}秒 | 行数: {ws.max_row: ,} , 列数: {ws.max_column: ,}"
        )

        self.logger.info(f"\n-- 转换 '{path.name}' --")

        chunk_size: int = 5_000
        max_row: int = ws.max_row
        max_col: int = ws.max_column
        total_chunks: int = (max_row + chunk_size - 1) // chunk_size

        with open(csv_path, "w", newline="", encoding=encoding) as f:
            writer = csv.writer(f)

            for chunk_idx, start_row in enumerate(
                tqdm(
                    range(1, max_row + 1, chunk_size),
                    total=total_chunks,
                    desc="转换进度",
                    unit="块",
                    ncols=100,
                    bar_format="{l_bar}{bar}{r_bar}",
                )
            ):
                end_row = min(start_row + chunk_size - 1, max_row)

                rows = ws.iter_rows(
                    min_row=start_row,
                    max_row=end_row,
                    min_col=1,
                    max_col=max_col,
                    values_only=True,
                )

                for row in rows:
                    writer.writerow(row)

        wb.close()

        self.logger.info(f"\n-- {path.name}转换完成 --")

        return csv_path

    def __dir_pattern(self) -> list[Path]:
        """dir模式

        Returns:
            list[Path]: csv文件路径
        """

        self.logger.info(f"\n--开始运行 dir 模式--")

        input_dir: Path = self.path_exists(
            input("请输入待转换的Excel文件夹路径:\t").strip().strip("'").strip('"')
        )
        output_dir: Path = self.path_exists(
            input("请输入保存Csv文件的文件夹路径: \t").strip().strip("'").strip('"')
        )

        csv_list: list[Path] = list()

        for p in input_dir.iterdir():
            csv_path: Path | None = self.to_csv(p, output_dir=output_dir)
            if csv_path:
                csv_list.append(csv_path)

        self.logger.info(f"成功转换{len(csv_list)}个文件.")

        return csv_list

    def __file_pattern(self) -> Path:
        """file模式

        Returns:
            Path: csv文件路径
        """

        self.logger.info(f"\n--开始运行 file 模式--")

        input_file: Path = self.path_exists(
            input("请输入待转换的Excel文件路径:\t").strip().strip("'").strip('"')
        )

        csv_path: Path | None = self.to_csv(input_file)

        if csv_path:
            return csv_path
        else:
            return input_file

    def operation(self) -> Path | list[Path] | None:
        """该类的主运行方法

        Returns:
            Path | list[Path]: 转换后的csv文件路径
        """

        self.logger.info("\n----ExcelToCsv模块-开始----")

        if self.method == "dir":
            result_list: list[Path] = self.__dir_pattern()
            self.logger.info("\n----ExcelToCsv模块-结束----")
            return result_list

        elif self.method == "file":
            result: Path = self.__file_pattern()
            self.logger.info("\n----ExcelToCsv模块-结束----")
            return result
